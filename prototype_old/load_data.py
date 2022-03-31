"""
The code for the ontology population with individual extracted
from tabular data file (Step 5 in Meteor technology).

Created: ???
Last change: 09 Apr 2021

@author: Irina Leshcheva
@contributor: Alena Begler
"""

from tkinter import Tk
from tkinter.filedialog import askopenfilename
from owlready2 import *
from os.path import abspath, dirname, join
from openpyxl import load_workbook
import types

# Получаем значения из столбца
def getRange(column):
    """
    column: ???.

    Returns data values from the corresponding column.
    """
    table = onto.search_one(hasPart = column)
    sheet = onto.search_one(hasPart = table)
    workbook = onto.search_one(hasPart = sheet)

    wb = load_workbook(filename = join(abspath(workbook.hasPath[0]),workbook.hasName[0]))
    ws = wb[sheet.label.first()]
    for tbl in ws.tables.values():
        if tbl.name == table.label.first():
            rng = ws[column.comment.first()]
    return rng

# Создаем дерево зависимостей для combineInstancesFrom, включая значения из столбцов
def makeTree(classList, assertion):
    """
    classList: a list of classes to be populated.
    assertion: ???.

    Returns a dictionary {'ontology class': 'corresponding values for instances creation'}
    and a length of the dictionary???
    """
    d = dict()
    fake = 0
    for item in classList:
        if item in assertion["getInstancesFrom"]:
            d[item] = getRange(item.getInstancesFrom[0])
            n = len(d[item])
        elif item in assertion["combineInstancesFrom"]:
            (d[item], fake) = makeTree(item.combineInstancesFrom, assertion)
    return (d, n)

def getLabel(dictRng, i):
    """
    dictRng : ???.
    i : ???.

    Returns ???.
    """
    label = ""
    for item in dictRng:
        if type(dictRng[item]) == dict:
            itemLabel = getLabel(dictRng[item], i)
        else:
            itemLabel = dictRng[item][i][0].value
        label = label + ":" + itemLabel            
    return label[1:]

def getLabels(columns, i):
    """
    dictRng : ???.
    i : ???.

    Returns dictionary ???.
    """
    d = dict()
    for klass in columns:
        if type(columns[klass]) == dict:
            d[klass] = getLabel(columns[klass], i)
        else:
            d[klass] = columns[klass][i][0].value
    return d       
   

def loadInstancesFrom(onto, dictClasses, domainOnto):
    """
    onto : ontology to be populated (filename???).
    The ontology should insclude assertions as annotation properties.
    dictClasses : ???.
    domainOnto : ???.

    Returns True while creating inctances using assrtion loadInstancesFrom.
    """
    # Список классов, экземпляры в которых должны быть объеденены в списки
    listMake = set(onto.search(makeList = "*"))
    ppointer = meteor.pointer
    # propSource = meteor.source
    
    # Проходим по всем классам, для которых необходимо создать экземпляры
    for entity in dictClasses:
        print("loadInstancesFrom " + entity.name)
                
        # Получаем имеющиеся экземпляры класса и кладем их в словарь
        classDict = dict()
        for inst in domainOnto[entity.name]["instances"]:
            print(inst)
            classDict[inst.label.first()] = inst
            if entity in listMake:
                if inst in ppointer.hasLast:
                    prev = inst
                    with onto:
                        ppointer.hasLast.remove(inst)
        print(classDict)
        # Получаем данные
        column = dictClasses[entity]
        table = onto.search_one(hasPart = column)
        sheet = onto.search_one(hasPart = table)
        workbook = onto.search_one(hasPart = sheet)

        wb = load_workbook(filename = join(abspath(workbook.hasPath[0]),workbook.hasName[0]))
        ws = wb[sheet.label.first()]
        for tbl in ws.tables.values():
            if tbl.name == table.label.first():
                rng = ws[column.comment.first()]

        # rng = getRange(assertion["getInstancesFrom"][entity])

        # Для каждой ячейки создаем экземпляр, если с таким label экземпляра еще нет
        n = len(classDict) # по ходу кода используется не n, a len(classDict)
        for cell in rng:
            cellValue = cell[0].value
            if cellValue!=  "" and cellValue not in classDict:
                instance = entity(namespace = onto, label = cellValue)
                # with onto:
                    # instance.propSource = workbook.hasName[0] + " | " + sheet.label[0] + " | " + cell[0].coordinate
                # print(workbook.hasName[0] + " | " + sheet.label[0] + " | " + cell[0].coordinate)
                classDict[cellValue] = instance
                
                # Если нужен список
                if entity in listMake:
                    if len(classDict) == 1: # если добавили первый элемент
                        with onto:
                            ppointer.hasFirst.append(instance) # говорим, что он первый
                    else: 
                        prev.hasNext = instance # связываем с предыдущим, если это не первый элемент
                    prev = instance

        # Если нужен список, то указываем на последний
        if entity in listMake:
            with onto:
                ppointer.hasLast.append(prev)
            
        domainOnto[entity.name]["dict"] = classDict
        domainOnto[entity.name]["filled"] = 1

    # print(domainOnto)
    # print(ppointer.hasFirst)
    # print(ppointer.hasLast)
    return True

def tryGetValuesFrom(onto, dictClasses, domainOnto, assertion):
    """
    onto : ontology to be populated (filename???).
    The ontology should insclude assertions as annotation properties.
    dictClasses : ???.
    domainOnto : ???.
    assertion : ???.

    Returns True while creating data properties using assrtion getValuesFrom.
    """
    for entity in dictClasses:
        print("tryGetValuesFrom " + entity.name)
        classDict = domainOnto[entity.name]["dict"]
        columnsData = dict()
        props = dict()
        for inst in dictClasses[entity]:
            rng = getRange(inst)
            columnsData[inst.label.first()] = rng
            # Узнаем свойство из аннотации useDataProperty
            props[inst.label.first()] = meteor.useDataProperty[entity, meteor.getValuesFrom, inst][0]
            n = len(rng)
        if entity in assertion["combineInstancesFrom"]:
            (columnsInst, fake) = makeTree(assertion['combineInstancesFrom'][entity], assertion)
        elif entity in assertion["getInstancesFrom"]:
            columnsInst = {entity: getRange(assertion["getInstancesFrom"][entity])}
        for i in range(n):
            label = getLabel(columnsInst, i)
            for name in columnsData:
                props[name][classDict[label]].append(columnsData[name][i][0].value)
                # meteor.source[classDict[label], props[name], columnsData[name][i][0].value] = 1
        domainOnto[entity.name]["filled"] = 2
    return True

def tryCombineInstancesFrom(onto, dictClasses, domainOnto, assertion):
    """
    onto : ontology to be populated (filename???).
    The ontology should insclude assertions as annotation properties.
    dictClasses : ???.
    domainOnto : ???.
    assertion : ???.

    Returns a number of instances created using assrtion combineInstancesFrom.
    """
    made = 0
    for entity in dictClasses:
        ready = 1
        for klass in dictClasses[entity]:
            ready = ready * domainOnto[klass.name]["filled"]
        if ready != 0 and domainOnto[entity.name]["filled"] == 0:
            # Получаем экземпляры класса и кладем их в словарь
            print("tryCombineInstancesFrom " + entity.name)
            classDict = dict()
            for inst in domainOnto[entity.name]["instances"]:
                classDict[inst.label.first()] = inst
            (columns, n) = makeTree(dictClasses[entity], assertion)
            for i in range(n):
                labels = getLabels(columns, i)
                label = ""
                for klass in dictClasses[entity]:
                    label = label + ":" + labels[klass]
                label = label[1:]
                if label not in classDict:
                    instance = entity(namespace = onto, label = label)
                    classDict[label] = instance
                    for klass in dictClasses[entity]:
                        # Узнаем свойство из аннотации useObjectProperty
                        prop = meteor.useObjectProperty[entity, meteor.combineInstancesFrom, klass][0]
                        prop[instance].append(domainOnto[klass.name]["dict"][labels[klass]])
            domainOnto[entity.name]["dict"] = classDict
            domainOnto[entity.name]["filled"] = 1
            made = made + 1
    return made

def tryMakeReferenceTo(onto, dictClasses, domainOnto, assertion):
    """
    onto : ontology to be populated (filename???).
    The ontology should insclude assertions as annotation properties.
    dictClasses : ???.
    domainOnto : ???.
    assertion : ???.

    Returns True while creating object properties using assrtion makeReferenceTo.
    """
    for entity in dictClasses:
        print("tryMakeReferenceTo " + entity.name)
        classDictFrom = domainOnto[entity.name]["dict"]
        (columns, n) = makeTree(dictClasses[entity], assertion)
        if entity in assertion["combineInstancesFrom"]:
            (columnsInst, fake) = makeTree(assertion["combineInstancesFrom"][entity], assertion)
        elif entity in assertion["getInstancesFrom"]:
            columnsInst = {entity: getRange(assertion["getInstancesFrom"][entity])}
        for i in range(n):
            labelFrom = getLabel(columnsInst, i)
            labelsTo = getLabels(columns, i)
            if labelFrom != "":
                for klass in dictClasses[entity]:
                    if labelsTo[klass] != "":
                        # Узнаем свойство из аннотации useObjectProperty
                        prop = meteor.useObjectProperty[entity, meteor.makeReferenceTo, klass][0]
                        classDictTo=domainOnto[klass.name]["dict"]
                        prop[classDictFrom[labelFrom]].append(classDictTo[labelsTo[klass]])
    return True

def loadDataToOntology(onto):
    """
    onto: ontology to be populated (filename???).
    The ontology should insclude assertions as annotation properties.

    Returns True while ontology populated with data from tabular file.
    """
    # Вспомогательный словарь, в котором хранится ссылка на класс, имеющиеся и созданные инстансы, а также индикатор, который показывает, заполнен ли класс новыми инстансами
    domainOnto = dict()
    for entity in list(onto.classes()):
        domainOnto[entity.name]={"class":entity, "instances":entity.direct_instances(), "filled":0}

    # Словарь, содержащий сущности и значения 4 типов аннотаций (getInstancesFrom, getValuesFrom, combineInstancesFrom, makeReferenceTo)
    assertion = dict()
   
    # Ищет классы, в аннотациях которых есть свойство getInstancesFrom
    assertion["getInstancesFrom"] = dict()
    for klass in onto.search(getInstancesFrom="*"):
        assertion["getInstancesFrom"][klass] = klass.getInstancesFrom[0]
        if klass.name not in domainOnto:
            domainOnto[klass.name] = {"class":klass, "instances":klass.direct_instances(), "filled":0}

    # Ищет классы, в аннотациях которых есть свойство combineInstancesFrom
    assertion["combineInstancesFrom"] = dict()
    for klass in onto.search(combineInstancesFrom = "*"):
        assertion["combineInstancesFrom"][klass] = klass.combineInstancesFrom
        if klass.name not in domainOnto:
            domainOnto[klass.name] = {"class":klass, "instances":klass.direct_instances(), "filled":0}
            
    # Ищет классы, в аннотациях которых есть свойство getValuesFrom
    assertion["getValuesFrom"] = dict()
    for klass in onto.search(getValuesFrom = "*"):
        assertion["getValuesFrom"][klass] = klass.getValuesFrom
        if klass.name not in domainOnto:
            domainOnto[klass.name] = {"class":klass, "instances":klass.direct_instances(), "filled":0}
            
    # Ищет классы, в аннотациях которых есть свойство makeReferenceTo
    assertion["makeReferenceTo"] = dict()
    for klass in onto.search(makeReferenceTo="*"):
        assertion["makeReferenceTo"][klass] = klass.makeReferenceTo
        if klass.name not in domainOnto:
            domainOnto[klass.name] = {"class":klass, "instances":klass.direct_instances(), "filled":0}

    print("domainOnto: ")
    print(domainOnto)            
    print("Assertion: ")
    print(assertion)
    
    # Заполняет классы инстансами (пока считает, что из одного столбца)
    loadInstancesFrom(onto, assertion["getInstancesFrom"], domainOnto)
    # Заполняет классы комбинированными инстансами (не умеет объединять их в список)
    test = len(assertion["combineInstancesFrom"])
    made = 0
    while test>made:
        made = made + tryCombineInstancesFrom(onto, assertion["combineInstancesFrom"], domainOnto, assertion)    
    # Для экземпляров этих классов задает значения свойств
    tryGetValuesFrom(onto, assertion["getValuesFrom"], domainOnto, assertion)
    # Устанавливает связи между инстансами
    tryMakeReferenceTo(onto, assertion["makeReferenceTo"], domainOnto, assertion)

    return True
	
Tk().withdraw()
filename = askopenfilename()
#print(filename)
onto_path.append(dirname(filename))
#print(dirname(filename))
#filename="ontomorf.owl"

onto = get_ontology(filename).load()
meteor = onto.get_namespace("http://meteor.ru/meteor#")
meteor_onto = get_ontology("http://meteor.ru/meteor.owl")

print(onto)

loadDataToOntology(onto)
onto.save()


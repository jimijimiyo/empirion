"""
The code for the tabular file data structure extraction
and its unsertion in the ontology (Step 3 in Meteor technology).

Created: ???
Last change: 09 Apr 2021

@author: Irina Leshcheva
@contributor: Alena Begler
"""

from tkinter import Tk
from tkinter.filedialog import  askopenfilename
from owlready2 import *
from os.path import abspath, dirname, join
from openpyxl import load_workbook

def readWorkbookStructure(fullName):
    """
    fullName: name of the tabular file including filepath 
    (escape the backslash with another backslash (\\)).
    This function is called by a putWorkbookStructureToOntology(onto),
    in this case fullName is created automatically.

    Returns nested dictionary with three levels: 
    {'worksheet': {'table': {'column': 'cells range'}}}.
    """
    wb = load_workbook(filename = fullName)
    wbDict = {}
    for ws in wb.worksheets:
        print(ws.title)
        print(ws.tables)
        wbDict[ws.title] = {}
        for tbl in ws.tables.values():
            wbDict[ws.title][tbl.name] = {}
            [begin, end] = tbl.ref.split(':')
            b = ws[begin]
            e = ws[end]
            rng = ws[begin:end]
            headers = tuple(c for c in rng[0])
            for col in tbl.tableColumns:
                for cell in headers:
                    if col.name == cell.value:
                        ref = cell.column_letter
                wbDict[ws.title][tbl.name][col.name] = ref + str(b.row + 1) + ":" + ref + str(e.row)
            print(wbDict)
    return wbDict

def putWorkbookStructureToOntology(onto):
    """    
    onto: ontology in which tabular data file structure should be inserted.
    To access an ontology its http address is needed.

    Returns an ontology populated with data structure individuals.
    """
    meteor = onto.get_namespace("http://meteor.ru/meteor#") # не совсем поняла, зачем делаем это, если уже сделали get_ontology()
    workbooksList = onto.search(type = meteor.Workbook)
    for workbook in workbooksList:
        print(workbook.hasPath)
        print(workbook.hasName)
        filename = join(abspath(workbook.hasPath.first()),workbook.hasName.first())
        print(filename)
        wbStruct = readWorkbookStructure(filename)
        workbook.hasPart = []
        for key1 in wbStruct:
            sheet = meteor.Sheet(namespace = onto, label = key1)
            workbook.hasPart.append(sheet)
            print(sheet.name)
            print(sheet.label)
            sheet.hasPart = []
            for key2 in wbStruct[key1]:
                table = meteor.Table(namespace = onto, label = key2)
                sheet.hasPart.append(table)
                table.hasPart = []
                for item in wbStruct[key1][key2]:
                    print(key1)
                    print(key2)
                    print(item)
                    col = meteor.Column(namespace = onto, label = item, comment = wbStruct[key1][key2][item])
                    table.hasPart.append(col)
    onto.save()
    return 1 # предлагаю что-то осмысленное, например "Workbook structure has been inserted in the ontology"
            # в идеале вообще с именем таблицы и онтологии, для которых это делалось
           

Tk().withdraw() # а это что делает?
filename = askopenfilename()
print(filename)
onto_path.append(dirname(filename))
print(dirname(filename))
onto = get_ontology("file://"+filename).load()

# может это закомментить? или лучше их показывать при каждом запуске?
# print(list(onto.classes()))
# print(list(onto.imported_ontologies))
# print(list(onto.individuals()))
# print(list(onto.object_properties()))
# print(list(onto.data_properties()))
# print(list(onto.annotation_properties()))

putWorkbookStructureToOntology(onto)




